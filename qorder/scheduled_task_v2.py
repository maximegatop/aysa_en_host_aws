import os
import sys
import django
from django.conf import settings
from openpyxl import load_workbook
from qorder.models import PuntoDeSuministro

# Seteamos la ruta del proyecto, para que se conozca el módulo bat_webservice_python 
project_path = '/home/ubuntu/aysaqorder/'
sys.path.append(project_path)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

def run_scheduled_task():
    try:
        file_path = '/home/ubuntu/soporte/gps_pds_20241031.xlsx'
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active

        puntos_suministros = []
        instancias_puntos_suministros = []
        puntos_suministro_model = {punto.punto_suministro: punto for punto in PuntoDeSuministro.objects.all()}

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                
                gps_1 = row[3]
                gps_2 = row[4]
                gps_3 = row[5]
                suministro = row[2]
                
                promedio_latitud = 0
                promedio_longitud = 0
                lista_latitud = []
                lista_longitud = []
                
                if gps_1 != '[NULL]':
                    latitud = float(gps_1.split(',')[0])
                    longitud = float(gps_1.split(',')[1])
                    lista_latitud.append(latitud)
                    lista_longitud.append(longitud)
                
                if gps_2 != '[NULL]':
                    latitud = float(gps_2.split(',')[0])
                    longitud = float(gps_2.split(',')[1])
                    lista_latitud.append(latitud)
                    lista_longitud.append(longitud)
                
                if gps_3 != '[NULL]':
                    latitud = float(gps_3.split(',')[0])
                    longitud = float(gps_3.split(',')[1])
                    lista_latitud.append(latitud)
                    lista_longitud.append(longitud)
                    
                if len(lista_latitud) > 0:
                    promedio_latitud = str(round(sum(lista_latitud) / len(lista_latitud), 7))
                    
                if len(lista_longitud) > 0:
                    promedio_longitud = str(round(sum(lista_longitud) / len(lista_longitud), 7))
                
                suministro = str(row[2])
                    
                puntos_suministros.append({'punto_suministro': suministro, 'promedio_latitud':promedio_latitud, 'promedio_longitud': promedio_longitud})

            except Exception as e:
                print(f'Error en fila: {e}')

        for punto in puntos_suministros:
            punto_suministro = punto['punto_suministro']
            punto_latitud = punto['promedio_latitud']
            punto_longitud = punto['promedio_longitud']

            try:
                if punto_suministro in puntos_suministro_model:
                    if punto_latitud != '0' and punto_longitud != '0':
                        puntos_suministro_model[punto_suministro].gps_latitud = punto_latitud
                        puntos_suministro_model[punto_suministro].gps_longitud = punto_longitud
                        puntos_suministro_model[punto_suministro].save()
            except Exception as e:
                print('Error en proceso de inserción')

        wb.close()
                
    except Exception as e:
        print(f'Error en proceso: {e}')

if __name__ == "__main__":
    run_scheduled_task()
